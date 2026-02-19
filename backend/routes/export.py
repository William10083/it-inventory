import pandas as pd
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import database, models
import io
from datetime import datetime

router = APIRouter()

def model_to_dict(models_list):
    data = []
    for m in models_list:
        d = m.__dict__.copy()
        if "_sa_instance_state" in d:
            del d["_sa_instance_state"]
        data.append(d)
    return data

@router.get("/export/excel")
def export_inventory_xlsx(db: Session = Depends(database.get_db)):
    # 1. Fetch Data
    devices = db.query(models.Device).all()
    employees = db.query(models.Employee).all()
    assignments = db.query(models.Assignment).all()
    maintenance = db.query(models.MaintenanceLog).all()

    # 2. Create DataFrames
    df_devices = pd.DataFrame(model_to_dict(devices))
    df_employees = pd.DataFrame(model_to_dict(employees))
    df_assignments = pd.DataFrame(model_to_dict(assignments))
    df_maintenance = pd.DataFrame(model_to_dict(maintenance))

    # 3. Write to Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if not df_devices.empty:
            df_devices.to_excel(writer, sheet_name='Devices', index=False)
        else:
            pd.DataFrame({'info': ['No devices']}).to_excel(writer, sheet_name='Devices')
            
        if not df_employees.empty:
            df_employees.to_excel(writer, sheet_name='Employees', index=False)
            
        if not df_assignments.empty:
            df_assignments.to_excel(writer, sheet_name='Assignments', index=False)
            
        if not df_maintenance.empty:
            df_maintenance.to_excel(writer, sheet_name='Maintenance', index=False)
            
    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="inventory_export.xlsx"'
    }
    return StreamingResponse(
        output, 
        headers=headers, 
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@router.get("/export/sales/excel")
def export_sales_xlsx(db: Session = Depends(database.get_db)):
    """
    Export all sales data to Excel with detailed information about each sale,
    including buyer info, devices sold, and prices.
    """
    from sqlalchemy.orm import joinedload
    
    # Fetch all sales with their items
    sales = db.query(models.Sale).options(
        joinedload(models.Sale.items),
        joinedload(models.Sale.sold_devices)
    ).all()
    
    # Prepare data for export
    sales_data = []
    
    for sale in sales:
        # Get devices info from sale_items (permanent record)
        devices_list = []
        total_devices = 0
        
        if sale.items:
            for item in sale.items:
                device_info = f"{item.device_type} {item.device_description}"
                if item.serial_number:
                    device_info += f" (S/N: {item.serial_number})"
                devices_list.append(device_info)
                total_devices += 1
        
        # Combine devices into a single string
        devices_str = " | ".join(devices_list) if devices_list else "N/A"
        
        sales_data.append({
            'ID Venta': sale.id,
            'Fecha de Venta': sale.sale_date.strftime('%d/%m/%Y %H:%M') if sale.sale_date else '',
            'Comprador': sale.buyer_name,
            'DNI': sale.buyer_dni,
            'Email': sale.buyer_email or '',
            'Teléfono': sale.buyer_phone or '',
            'Dirección': sale.buyer_address or '',
            'Cantidad de Dispositivos': total_devices,
            'Dispositivos Vendidos': devices_str,
            'Precio Total': sale.sale_price or 0,
            'Método de Pago': sale.payment_method or '',
            'Notas': sale.notes or '',
            'Tiene Acta': 'Sí' if sale.acta_path else 'No',
            'Creado Por': sale.created_by_user_id or '',
            'Fecha de Creación': sale.created_at.strftime('%d/%m/%Y %H:%M') if sale.created_at else ''
        })
    
    # Create DataFrame
    df_sales = pd.DataFrame(sales_data)
    
    # Create detailed items sheet
    items_data = []
    for sale in sales:
        if sale.items:
            for item in sale.items:
                items_data.append({
                    'ID Venta': sale.id,
                    'Fecha Venta': sale.sale_date.strftime('%d/%m/%Y') if sale.sale_date else '',
                    'Comprador': sale.buyer_name,
                    'Tipo Dispositivo': item.device_type,
                    'Descripción': item.device_description,
                    'Número de Serie': item.serial_number or 'N/A',
                    'Precio': item.price
                })
    
    df_items = pd.DataFrame(items_data)
    
    # Write to Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if not df_sales.empty:
            df_sales.to_excel(writer, sheet_name='Ventas', index=False)
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Ventas']
            for idx, col in enumerate(df_sales.columns):
                max_length = max(
                    df_sales[col].astype(str).apply(len).max(),
                    len(col)
                )
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 2, 50)
        else:
            pd.DataFrame({'info': ['No hay ventas registradas']}).to_excel(writer, sheet_name='Ventas')
        
        if not df_items.empty:
            df_items.to_excel(writer, sheet_name='Detalle Items', index=False)
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Detalle Items']
            for idx, col in enumerate(df_items.columns):
                max_length = max(
                    df_items[col].astype(str).apply(len).max(),
                    len(col)
                )
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 2, 50)
    
    output.seek(0)
    
    # Generate filename with current date
    today = datetime.now().strftime('%Y%m%d')
    filename = f"ventas_export_{today}.xlsx"
    
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    
    return StreamingResponse(
        output,
        headers=headers,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@router.get("/export/assignments-template")
def export_assignments_template(db: Session = Depends(database.get_db)):
    """
    Export detailed assignments report using the official template
    """
    from sqlalchemy.orm import joinedload
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side
    import os
    import json
    
    # Path to template
    import os
    
    # Calculate absolute path to template
    # export.py is in backend/routes/
    current_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(os.path.dirname(current_dir)) # Up 2 levels to it_inventory/
    template_path = os.path.join(project_root, "resources", "templates", "Inventario TI (DIC 2025) (1).xlsx")
    
    if not os.path.exists(template_path):
        # Raise error instead of fallback so we know if path is wrong
        raise HTTPException(status_code=500, detail=f"Template not found at: {template_path}")
        
    # Load workbook
    # Load workbook
    wb = openpyxl.load_workbook(template_path)
    
    # Robust Sheet Selection
    target_name_lower = "inventario ti"
    found_sheet_name = None
    
    # 1. Try exact match (most reliable)
    if "Inventario TI" in wb.sheetnames:
        found_sheet_name = "Inventario TI"
    else:
        # 2. Try case-insensitive/stripped match
        for name in wb.sheetnames:
            if name.strip().lower() == target_name_lower:
                found_sheet_name = name
                break
        # 3. Fallback to index 1 (Second sheet as observed)
        if not found_sheet_name and len(wb.sheetnames) > 1:
            found_sheet_name = wb.sheetnames[1]
            
    if found_sheet_name:
        ws = wb[found_sheet_name]
        # Remove ALL other sheets safely
        # Use list(wb.sheetnames) to iterate over a copy while modifying the workbook
        for sheet_name in list(wb.sheetnames):
            if sheet_name != found_sheet_name:
                wb.remove(wb[sheet_name])
    else:
        # Last resort: use active sheet and keep assumed structure
        ws = wb.active
        
    # Fetch all employees (active)
    employees = db.query(models.Employee).filter(models.Employee.is_active == True).all()
    
    # Fetch all active assignments (not returned)
    assignments = db.query(models.Assignment).filter(
        models.Assignment.returned_date == None
    ).options(
        joinedload(models.Assignment.device),
        joinedload(models.Assignment.employee)
    ).all()
    
    # Group assignments by Employee ID
    emp_map = {e.id: {"employee": e, "devices": []} for e in employees}
    
    # Fill map with assignments
    for a in assignments:
        if a.employee_id in emp_map:
            emp_map[a.employee_id]["devices"].append(a.device)
            
            
    # Find header row dynamically to match template logic exactly
    start_row = 5 # default fallback
    header_found = False
    
    # Scan first 20 rows for "ITEM" header
    for row in ws.iter_rows(min_row=1, max_row=20):
        for cell in row:
            if cell.value == "ITEM":
                start_row = cell.row + 1
                header_found = True
                break
        if header_found:
            break
            
    # Clear existing data
    # Clear from start_row down to 1000 to ensure no dummy data remains
    for row in ws.iter_rows(min_row=start_row, max_row=1000):
        for cell in row:
            cell.value = None
            
    row_idx = start_row
    item_counter = 1
    
    sorted_employees = sorted(employees, key=lambda x: x.full_name or "")
    
    for emp in sorted_employees:
        data = emp_map.get(emp.id, {"devices": []})
        devices = data["devices"]
        
        # Skip employees with no assignments? 
        # User said "Inventario", so maybe list everyone? 
        # But usually you only list people with equipment.
        # Let's list everyone to be safe, or just those with assignments. 
        # If I look at the template name "Inventario TI", it implies inventory.
        # But "Asignaciones" implies active ones.
        # I'll include everyone who has at least one assignment for now, to keep it clean.
        if not devices:
            continue
            
        # Classify Devices
        laptop = next((d for d in devices if d.device_type in ['laptop', 'pc', 'desktop', 'all-in-one']), None)
        monitor = next((d for d in devices if d.device_type == 'monitor'), None)
        keyboard = next((d for d in devices if d.device_type == 'keyboard'), None)
        mouse = next((d for d in devices if d.device_type == 'mouse'), None)
        charger = next((d for d in devices if d.device_type == 'charger'), None) # Explicit assignment
        backpack = next((d for d in devices if d.device_type == 'backpack' or 'mochila' in (d.device_type or "").lower()), None)
        kit = next((d for d in devices if d.device_type and d.device_type.lower() in ['kit teclado/mouse', 'kit']), None)
        
        # Refined Headset detection (Computer headsets)
        # Assuming 'auriculares' or 'headset' or specific models
        headset = next((d for d in devices if d.device_type and d.device_type.lower() in ['headset', 'auriculares', 'headphones', 'diadema']), None)
        
        # Parse specifications if available
        specs = {}
        if laptop and laptop.specifications:
            try:
                if isinstance(laptop.specifications, str):
                    specs = json.loads(laptop.specifications)
                elif isinstance(laptop.specifications, dict):
                    specs = laptop.specifications
            except:
                pass
                
        # --- Columns Mapping (1-based) ---
        # 1: ITEM
        ws.cell(row=row_idx, column=1, value=item_counter)
        
        # 2: EMPRESA
        ws.cell(row=row_idx, column=2, value=emp.company or "")
        
        # 3: DNI
        ws.cell(row=row_idx, column=3, value=emp.dni)
        
        # 4: NOMBRES / 5: APELLIDOS (Split full name)
        full_name = emp.full_name or ""
        parts = full_name.split()
        if len(parts) >= 3: # e.g. "Juan Carlos Perez"
            ws.cell(row=row_idx, column=4, value=" ".join(parts[:2])) # Nombres
            ws.cell(row=row_idx, column=5, value=" ".join(parts[2:])) # Apellidos
        elif len(parts) == 2:
            ws.cell(row=row_idx, column=4, value=parts[0])
            ws.cell(row=row_idx, column=5, value=parts[1])
        else:
             ws.cell(row=row_idx, column=4, value=full_name)
             
        # 6: SEDE
        ws.cell(row=row_idx, column=6, value=emp.location)
        
        # 7: AREA
        ws.cell(row=row_idx, column=7, value=emp.department)
        
        # 8: PUESTO
        ws.cell(row=row_idx, column=8, value=emp.position)
        
        # --- DATOS DEL EQUIPO (Laptop/PC) ---
        if laptop:
            ws.cell(row=row_idx, column=9, value=laptop.device_type.upper() or "NA") # TIPO DE EQUIPO
            ws.cell(row=row_idx, column=10, value="ASIGNACION") # TIPO DE ENTREGA
            ws.cell(row=row_idx, column=11, value=laptop.hostname or "NA") # NOMBRE DEL EQUIPO
            ws.cell(row=row_idx, column=12, value=laptop.inventory_code or "NA") # CODIGO INTERNO
            ws.cell(row=row_idx, column=13, value=laptop.brand or "NA") # MARCA
            ws.cell(row=row_idx, column=14, value=laptop.model or "NA") # MODELO
            ws.cell(row=row_idx, column=15, value=laptop.serial_number or "NA") # SERIE
            
            # Specs
            processor = specs.get('processor')
            storage = specs.get('storage')
            ram = specs.get('ram')
            
            # 2026-02-17 Update: Force specific specs for ProBook 440 and EliteBook as requested
            if laptop.model:
                model_lower = laptop.model.lower()
                if "440" in model_lower or "elitebook" in model_lower:
                    processor = "Intel(R) Core(TM) Ultra 7 155U (1.70 GHz)"
                    storage = "1TB SSS"
                    ram = "16.0 GB"
            
            ws.cell(row=row_idx, column=16, value=processor or "NA") # PROCESADOR
            ws.cell(row=row_idx, column=17, value=storage or "NA") # DISCO
            ws.cell(row=row_idx, column=18, value=ram or "NA") # RAM
            
        # --- CARGADOR (Cols 19-21) ---
        if charger:
             ws.cell(row=row_idx, column=19, value=charger.brand or "NA")
             ws.cell(row=row_idx, column=20, value=charger.model or "NA")
             ws.cell(row=row_idx, column=21, value="") # Requested to be blank
        elif laptop:
            # Maybe charger is embedded in specs?
            pass
            
        # --- TECLADO (Cols 22-25) ---
        # Priority: Specific Keyboard -> Kit -> None
        if keyboard:
            ws.cell(row=row_idx, column=22, value=keyboard.brand or "NA")
            ws.cell(row=row_idx, column=23, value=keyboard.model or "NA")
            ws.cell(row=row_idx, column=24, value=keyboard.serial_number or "NA")
            ws.cell(row=row_idx, column=25, value=keyboard.inventory_code or "NA")
        elif kit:
            # Use Kit details for Keyboard
            ws.cell(row=row_idx, column=22, value=kit.brand or "NA")
            ws.cell(row=row_idx, column=23, value="HSA-A005K") # Forced Model for Kit Keyboard works
            ws.cell(row=row_idx, column=24, value=kit.serial_number or "NA")
            ws.cell(row=row_idx, column=25, value=kit.inventory_code or "NA")
            
        # --- MOUSE (Cols 26-28) ---
        # Priority: Specific Mouse -> Kit -> None
        if mouse:
            ws.cell(row=row_idx, column=26, value=mouse.brand or "NA")
            ws.cell(row=row_idx, column=27, value=mouse.model or "NA")
            ws.cell(row=row_idx, column=28, value=mouse.serial_number or "NA")
        elif kit:
            # Use Kit details for Mouse
            ws.cell(row=row_idx, column=26, value=kit.brand or "NA")
            ws.cell(row=row_idx, column=27, value="HSA-A011M") # Forced Model for Kit Mouse works
            ws.cell(row=row_idx, column=28, value=kit.serial_number or "NA")
            
        # --- MONITOR (Cols 29-32) ---
        if monitor:
             ws.cell(row=row_idx, column=29, value=monitor.brand or "NA")
             ws.cell(row=row_idx, column=30, value=monitor.model or "NA")
             ws.cell(row=row_idx, column=31, value=monitor.serial_number or "NA")
             ws.cell(row=row_idx, column=32, value=monitor.inventory_code or "NA")
             
        # --- MOCHILA (Cols 33-36) ---
        if backpack:
             ws.cell(row=row_idx, column=33, value=backpack.brand or "NA")
             ws.cell(row=row_idx, column=34, value=backpack.model or "NA")
             ws.cell(row=row_idx, column=35, value=backpack.serial_number or "NA")
             
        # --- HEADSETS (Cols 37-40) ---
        if headset:
            ws.cell(row=row_idx, column=37, value=headset.brand or "NA")
            ws.cell(row=row_idx, column=38, value=headset.model or "NA")
            ws.cell(row=row_idx, column=39, value=headset.serial_number or "NA")
            
        
        # --- OTROS ACCESORIOS (Cols 45-49) ---
        # Logic: Only populate if person has a SECOND laptop assigned.
        # Find secondary laptop from the list we filtered earlier
        # We need to re-find laptops because 'laptop' variable above is just the first one found by next()
        
        all_laptops = [d for d in devices if d.device_type and d.device_type.lower() in ['laptop', 'pc', 'desktop', 'all-in-one']]
        
        if len(all_laptops) > 1:
            # The first one was used for main columns (9-18).
            # The second one goes here.
            secondary_laptop = all_laptops[1]
            
            # 45: TIPO
            ws.cell(row=row_idx, column=45, value=secondary_laptop.device_type.upper() or "NA")
            # 46: MARCA
            ws.cell(row=row_idx, column=46, value=secondary_laptop.brand or "NA")
            # 47: MODELO
            ws.cell(row=row_idx, column=47, value=secondary_laptop.model or "NA")
            # 48: SERIE
            ws.cell(row=row_idx, column=48, value=secondary_laptop.serial_number or "NA")
            # 49: CODIGO INTERNO
            ws.cell(row=row_idx, column=49, value=secondary_laptop.inventory_code or "NA")
            
        row_idx += 1
        item_counter += 1
        
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    today = datetime.now().strftime('%Y%m%d')
    filename = f"Inventario_TI_Reporte_{today}.xlsx"
    
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    
    return StreamingResponse(
        output,
        headers=headers,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
